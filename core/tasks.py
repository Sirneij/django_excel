import logging
import time
from datetime import datetime
from io import BytesIO
from typing import Any, Generator, Optional

import requests
from celery import shared_task
from decouple import config
from django.conf import settings
from django.core.mail import EmailMessage
from django.utils import timezone
from google.oauth2 import service_account
from googleapiclient.discovery import build
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Protection

from core.models import Coins, FullCoin
from core.templatetags.custom_tags import currency

logger = logging.getLogger(__name__)


@shared_task
def get_coins_data_from_coingecko_and_store() -> None:
    """Fetch data from coingecko api and store."""
    market_currency_order = 'markets?vs_currency=ngn&order=market_cap_desc&'
    per_page = 'per_page=250&page=1&sparkline=false'
    final_url = f'{settings.BASE_API_URL}{market_currency_order}{per_page}'

    coin_data = requests.get(final_url).json()

    for data in coin_data:
        coin, _ = Coins.objects.get_or_create(name=data['name'], symbol=data['symbol'])
        coin.image_url = data['image']
        coin.current_price = data['current_price']
        coin.price_change_within_24_hours = data['price_change_24h']
        coin.rank = data['market_cap_rank']
        coin.market_cap = data['market_cap']
        coin.total_supply = data['total_supply']
        coin.save()


@shared_task
def export_data_to_excel(user_email: str) -> None:
    """Send extracted model data and save in excel and send to email."""
    excelfile = BytesIO()
    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet = workbook.create_sheet(title='Latest Cryptocurrency Coins', index=1)
    workbook.security.workbookPassword = config('PASSWORD', default='12345data')
    workbook.security.lockStructure = config('PROTECT', default=True, cast=bool)
    workbook.security.revisionsPassword = config('PASSWORD', default='12345data')
    worksheet.protection.sheet = config('PROTECT', default=True, cast=bool)
    worksheet.protection.formatCells = config('PROTECT', default=False, cast=bool)

    worksheet.sheet_properties.tabColor = '1072BA'
    worksheet.freeze_panes = 'I2'

    coin_queryset = Coins.objects.all().order_by('rank')
    columns = ['Name', 'Symbol', 'Rank', 'Current price', 'Price change', 'Market cap', 'Total supply']
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.font = Font(bold=True)
    # Iterate through all coins
    for _, coin in enumerate(coin_queryset, 1):
        row_num += 1

        # Define the data for each cell in the row
        row = [
            coin.name,
            f'{coin.symbol}'.upper(),
            coin.rank,
            currency(coin.current_price),
            currency(coin.price_change_within_24_hours),
            currency(coin.market_cap),
            coin.total_supply,
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value
            cell.protection = Protection(locked=True)
    workbook.save(excelfile)
    now = timezone.now()
    message = EmailMessage(
        f'Coin data as of {now.date().isoformat()}',
        f'Generated at: {now.isoformat()}',
        settings.DEFAULT_FROM_EMAIL,
        [user_email],
    )
    message.attach('latest-coin-list.xlsx', excelfile.getvalue(), 'application/vnd.ms-excel')
    message.send()


@shared_task
def populate_googlesheet_with_coins_data() -> None:
    """Populate Googlesheet with the coin data from the database."""
    response = requests.get(settings.GOOGLE_API_SERVICE_KEY_URL)
    with open('core/djangoexcel.json', 'wb') as file:
        file.write(response.content)
    service_account_file = 'core/djangoexcel.json'
    creds = service_account.Credentials.from_service_account_file(
        service_account_file, scopes=settings.GOOGLE_API_SCOPE
    )
    service = build('sheets', 'v4', credentials=creds)
    sheet = service.spreadsheets()

    sheet_metadata_values = sheet.get(spreadsheetId=settings.SPREADSHEET_ID).execute()
    csheets = sheet_metadata_values.get('sheets', '')
    datetime_format = '%a %b %d %Y %Hh%Mm'
    if csheets and len(csheets) > 1:
        for csheet in csheets:
            sheet_title = csheet.get('properties', {}).get('title', '')
            date_segment_of_the_title = ' '.join(sheet_title.split(' ')[0:5]).strip()
            parsed_datetime: Optional[Any] = None
            try:
                parsed_datetime = datetime.strptime(date_segment_of_the_title, datetime_format)
            except ValueError as err:
                print(err)
            now = timezone.now().strftime(datetime_format)
            if (
                parsed_datetime
                and (datetime.strptime(now, datetime_format) - parsed_datetime).seconds
                > settings.SPREADSHEET_TAB_EXPIRY
            ):
                sheet_id = csheet.get('properties', {}).get('sheetId', 0)
                batch_update_request_body = {'requests': [{'deleteSheet': {'sheetId': sheet_id}}]}
                sheet.batchUpdate(spreadsheetId=settings.SPREADSHEET_ID, body=batch_update_request_body).execute()

    coin_queryset = Coins.objects.all().order_by('rank')
    coin_data_list: list[Any] = [
        [
            'Name',
            'Symbol',
            'Rank',
            'Current price',
            'Price change',
            'Market cap',
            'Total supply',
        ]
    ]
    for coin in coin_queryset:
        coin_data_list.append(
            [
                coin.name,
                f'{coin.symbol}'.upper(),
                coin.rank,
                str(currency(coin.current_price)),
                str(currency(coin.price_change_within_24_hours)),
                str(currency(coin.market_cap)),
                str(coin.total_supply),
            ]
        )

    new_sheet_title = f'{timezone.now().strftime(datetime_format)} Coin data'
    batch_update_request_body = {
        'requests': [
            {
                'addSheet': {
                    'properties': {
                        'title': new_sheet_title,
                        'tabColor': {'red': 0.968627451, 'green': 0.576470588, 'blue': 0.101960784},
                        'gridProperties': {'rowCount': len(coin_data_list), 'columnCount': 7},
                    }
                }
            }
        ]
    }
    sheet.batchUpdate(spreadsheetId=settings.SPREADSHEET_ID, body=batch_update_request_body).execute()
    sheet.values().append(
        spreadsheetId=settings.SPREADSHEET_ID,
        range=f"'{new_sheet_title}'!A1:G1",
        valueInputOption='USER_ENTERED',
        body={'values': coin_data_list},
    ).execute()


def build_api_url(page: int) -> str:
    """Build the API URL."""
    market_currency_order = 'markets?vs_currency=usd&order=market_cap_desc&'
    per_page = f'per_page=50&page={page}&sparkline=false'
    return f'{settings.BASE_API_URL}/coins/{market_currency_order}{per_page}'


def store_data(data_list: list[dict]) -> None:
    """Store the data in bulk."""
    # Get existing coins
    existing_coins = {coin.coin_id: coin for coin in FullCoin.objects.filter(coin_id__in=[d['id'] for d in data_list])}

    coins_to_create = []
    coins_to_update = []

    for data in data_list:
        if data['id'] in existing_coins:
            # Update existing coin
            coin = existing_coins[data['id']]
            coins_to_update.append(coin)
        else:
            # Create new coin
            coin = FullCoin(
                coin_id=data['id'],
                symbol=data['symbol'],
                name=data['name'],
                image=data['image'],
                current_price=data['current_price'],
                market_cap=data['market_cap'],
                market_cap_rank=data['market_cap_rank'],
                fully_diluted_valuation=data.get('fully_diluted_valuation'),
                total_volume=data['total_volume'],
                high_24h=data['high_24h'],
                low_24h=data['low_24h'],
                price_change_24h=data['price_change_24h'],
                price_change_percentage_24h=data['price_change_percentage_24h'],
                market_cap_change_24h=data['market_cap_change_24h'],
                market_cap_change_percentage_24h=data['market_cap_change_percentage_24h'],
                circulating_supply=data['circulating_supply'],
                total_supply=data.get('total_supply'),
                max_supply=data.get('max_supply'),
                ath=data['ath'],
                ath_change_percentage=data['ath_change_percentage'],
                ath_date=data['ath_date'],
                atl=data['atl'],
                atl_change_percentage=data['atl_change_percentage'],
                atl_date=data['atl_date'],
                last_updated=data['last_updated'],
            )
            coins_to_create.append(coin)
            continue

        # Update fields for existing coins
        coin.symbol = data['symbol']
        coin.name = data['name']
        coin.image = data['image']
        coin.current_price = data['current_price']
        coin.market_cap = data['market_cap']
        coin.market_cap_rank = data['market_cap_rank']
        coin.fully_diluted_valuation = data.get('fully_diluted_valuation')  # Using get() for nullable fields
        coin.total_volume = data['total_volume']
        coin.high_24h = data['high_24h']
        coin.low_24h = data['low_24h']
        coin.price_change_24h = data['price_change_24h']
        coin.price_change_percentage_24h = data['price_change_percentage_24h']
        coin.market_cap_change_24h = data['market_cap_change_24h']
        coin.market_cap_change_percentage_24h = data['market_cap_change_percentage_24h']
        coin.circulating_supply = data['circulating_supply']
        coin.total_supply = data.get('total_supply')  # Using get() for nullable fields
        coin.max_supply = data.get('max_supply')  # Using get() for nullable fields
        coin.ath = data['ath']
        coin.ath_change_percentage = data['ath_change_percentage']
        coin.ath_date = data['ath_date']
        coin.atl = data['atl']
        coin.atl_change_percentage = data['atl_change_percentage']
        coin.atl_date = data['atl_date']
        coin.last_updated = data['last_updated']

    # Bulk create new coins with ignore_conflicts=True
    if coins_to_create:
        FullCoin.objects.bulk_create(coins_to_create, ignore_conflicts=True)

    # Bulk update existing coins
    if coins_to_update:
        FullCoin.objects.bulk_update(
            coins_to_update,
            fields=[
                'symbol',
                'name',
                'image',
                'current_price',
                'market_cap',
                'market_cap_rank',
                'fully_diluted_valuation',
                'total_volume',
                'high_24h',
                'low_24h',
                'price_change_24h',
                'price_change_percentage_24h',
                'market_cap_change_24h',
                'market_cap_change_percentage_24h',
                'circulating_supply',
                'total_supply',
                'max_supply',
                'ath',
                'ath_change_percentage',
                'ath_date',
                'atl',
                'atl_change_percentage',
                'atl_date',
                'last_updated',
            ],
        )


def fetch_coins_iteratively() -> Generator[dict, None, None]:
    """Fetch coins data from API using generator."""
    page = 1
    while True:
        try:
            url = build_api_url(page)
            response = requests.get(url)
            coin_data = response.json()

            # Check for rate limit response
            if isinstance(coin_data, dict) and coin_data.get('status', {}).get('error_code') == 429:
                logger.warning("Rate limit exceeded. Waiting 60 seconds...")
                time.sleep(60)
                continue

            # Check for empty response (end of pagination)
            if not coin_data:
                break

            yield from coin_data
            logger.info(f"Fetched page {page} with {len(coin_data)} coins")
            page += 1

            time.sleep(1)  # Be nice to the API

        except requests.exceptions.RequestException as e:
            logger.error(f"Request failed on page {page}: {e}")
            raise


def fetch_coins_recursively(page: int = 1) -> Generator[dict, None, None]:
    """Fetch coins data from API recursively using generator."""
    try:
        url = build_api_url(page)
        response = requests.get(url)
        coin_data = response.json()

        # Check for rate limit response
        if isinstance(coin_data, dict) and coin_data.get('status', {}).get('error_code') == 429:
            logger.warning("Rate limit exceeded. Waiting 60 seconds...")
            time.sleep(60)
            yield from fetch_coins_recursively(page)
            return

        # Base case: empty response (end of pagination)
        if not coin_data:
            return

        # Process current page
        yield from coin_data
        logger.info(f"Fetched page {page} with {len(coin_data)} coins")

        # Be nice to the API
        time.sleep(1)

        # Recursive case: fetch next page
        yield from fetch_coins_recursively(page + 1)

    except requests.exceptions.RequestException as e:
        logger.error(f"Request failed on page {page}: {e}")
        raise


@shared_task(
    bind=True,
    autoretry_for=(Exception,),
    retry_backoff=True,
    retry_backoff_max=600,
    max_retries=5,
)
def get_full_coin_data_iteratively_for_page(self) -> None:
    """Get full coin data iteratively for each page."""
    try:
        # Use list comprehension to collect coins in batches
        batch_size = 100
        coins_batch = []

        for coin in fetch_coins_iteratively():
            coins_batch.append(coin)

            if len(coins_batch) >= batch_size:
                logger.info(f"Processing batch of {len(coins_batch)} coins")
                store_data(coins_batch)
                coins_batch = []

        # Process remaining coins
        if coins_batch:
            logger.info(f"Processing final batch of {len(coins_batch)} coins")
            store_data(coins_batch)

    except Exception as e:
        logger.error(f"Failed to process coins: {e}")
        raise self.retry(exc=e)
